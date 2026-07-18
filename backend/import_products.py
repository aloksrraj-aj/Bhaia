from openpyxl import load_workbook

from app.database.connection import SessionLocal
from app.models.category import Category
from app.models.product import Product


db = SessionLocal()

workbook = load_workbook("Bhaia_Products_50.xlsx")
sheet = workbook.active

category_map = {}

# Create categories
for row in sheet.iter_rows(min_row=2, values_only=True):

    name, brand, image, unit, category = row

    existing = (
        db.query(Category)
        .filter(Category.name == category)
        .first()
    )

    if not existing:

        existing = Category(name=category)

        db.add(existing)
        db.commit()
        db.refresh(existing)

    category_map[category] = existing.id


# Create products
for row in sheet.iter_rows(min_row=2, values_only=True):

    name, brand, image, unit, category = row

    already_exists = (
        db.query(Product)
        .filter(Product.name == name)
        .first()
    )

    if already_exists:
        continue

    product = Product(
        name=name,
        brand=brand,
        image=image,
        unit=unit,
        category_id=category_map[category]
    )

    db.add(product)

db.commit()

print("===================================")
print("Products Imported Successfully!")
print("===================================")

db.close()